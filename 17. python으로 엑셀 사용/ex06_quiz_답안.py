"""

    제주모법업소.xlsx 파일의 모든 업소들을
    '업태'별로 분류하여 새 파일들을 생성

    파일명 예) 탕류.xlsx, 중식.xlsx, 복어취급.xlsx ...

    파일에 업소명, 주소만 저장한다.

"""

import openpyxl

filename = '제주모범업소.xlsx'
wb = openpyxl.load_workbook(filename)
ws = wb['목록1']

category = {}

for i in range(3, 297):
    a = ws.cell(i, 5).value  # i번째 행의 E열 --> 업태
    a = a.replace('/', '_')  # '/'를 '_'로 변경 (파일명, sheet명은 '/'안됨..?)
    b = ws.cell(i, 2).value  # i번째 행의 B열 --> 업소명
    c = ws.cell(i, 4).value  # i번째 행의 D열 --> 주소
    if a not in category:
        category[a] = []
    category[a].append([b, c])

new_wb = openpyxl.Workbook()
for k in category:
    ws = new_wb.create_sheet(f'{k}.xlsx')
    lst = category[k]
    for i in range(len(lst)):  # for e in lst:
        ws.append(lst[i])      #    ws.append(e)

new_wb.save('업태별_제주모범업소.xlsx')