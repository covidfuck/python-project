"""

pip install openpyxl
"""
import openpyxl

# 작업할 엑셀 파일 생성 ( 진짜로 생성되지는 않은 상태 )
wb = openpyxl.Workbook()

# 기본 Sheet 불러오기
ws = wb.active

# B열 2행에 'abc' 저장
ws['B2'] = 'abc'

# 마지막 행에 줄 추가
ws.append(['a', 'b', 'c'])
ws.append(['가', '나', '다'])

# 특정 셀 접근
ws.cell(1, 2, 'book') # 1행 2열에 'book' 저장

# Workbook 저장 (경로 명시하지 않을시 현재 디렉토리에 저장) -> wb.save('ex01.xlsx')
# wb.save("C:\\Users\\정동연\\Desktop\\ex01.xlsx") # 경로 명시하지 않을시 현재 디렉토리에 저장
wb.save('ex01.xlsx')
