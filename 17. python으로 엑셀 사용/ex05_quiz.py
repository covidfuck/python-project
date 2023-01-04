"""

    제주모법업소.xlsx 파일의 모든 업소들을
    '업태'별로 분류하여 새 파일들을 생성

    파일명 예) 탕류.xlsx, 중식.xlsx, 복어취급.xlsx ...

    파일에 업소명, 주소만 저장한다.

"""
import openpyxl

filename = '제주모범업소.xlsx'
wb = openpyxl.load_workbook(filename)
ws = wb['목록1'] # Sheet 이름으로 불러오기

category = {}

for i in range(3, 297):
    a = ws.cell(i, 5).value
    name = ws.cell(i, 2).value
    address = ws.cell(i, 4).value
    lst = []
    lst.append(name)
    lst.append(address)
    if a in category:
        pass
    else:
        category[a] = []
    category[a].append(lst)

print(category)
b = list(category.keys())

for i in range(len(b)):
    wb = openpyxl.Workbook()

    ws = wb.create_sheet(f'{b[i]}')

    ws.append(['업소명', '소재지(도로명)'])

    for j in range(len(category[b[i]])):
        ws.append(category[b[i]][j])

    wb.save(f'{b[i]}.xlsx')








