import openpyxl

filename = 'quiz.xlsx'
wb = openpyxl.load_workbook(filename)

# Sheet 이름으로 불러오기
ws = wb['Sheet']

# 방법 1
a = ws['A1'].value
print(a)

# 방법 2
a = ws.cell(1, 2).value
print(a)

# 방법 3
a = ws.rows
for row in ws.rows: # '행'들로 반복 처리
    for col in row: # 각 '행'의 '열'들을 반복 처리 ==>
        print(col.value)
